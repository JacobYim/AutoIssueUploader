from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import chromedriver_autoinstaller
from selenium.webdriver.support.ui import Select
from utils import * 
import pandas as pd
import time
import tkinter as tk
import threading
import os
from openpyxl import load_workbook



def type_retriever(driver, issue_type = 'Bug') :

    if issue_type == 'Bug' :
        tab = driver.find_element(By.ID, "tab-0")
    else :
        tab = driver.find_element(By.CLASS_NAME, "content")
    children = tab.find_elements(By.XPATH, "*")

    retval = {}

    for idx_child, child in enumerate(children) :

        # initialize inputs
        field_name = None
        input_type = ""
        xpath = None
        required = False

        # check input element is required or not
        if  len(child.find_elements(By.CLASS_NAME, 'icon-required')) > 0 :
            required = True
        else :
            required = False
        
        child_class = child.get_attribute('class')
        if 'field-group' in child_class :
            # find field name
            input_type += 'field-group'
            field_name = child.find_element(By.TAG_NAME, 'label').get_attribute("innerText").split('\n')[0]

            dynamicselects = child.find_elements(By.CLASS_NAME, 'dynamic-select-mark')
            jiramultiselects = child.find_elements(By.CLASS_NAME, 'jira-multi-select')
            cfselects = child.find_elements(By.CLASS_NAME, 'cf-select')
            auissselects  = child.find_elements(By.CLASS_NAME, 'aui-ss-select')
            textareas = child.find_elements(By.CLASS_NAME, 'textarea')
            selects = child.find_elements(By.CLASS_NAME, 'select')
            texts = child.find_elements(By.CLASS_NAME, 'text')
            
            if 'aui-field-cascadingselect' in child_class :
                input_type += " aui-field-cascadingselect"

                retval[field_name] = {
                    'input_type' : input_type+" cascadingselect-parent", 
                    'required' : required,
                    'element' : child.find_element(By.CLASS_NAME, 'cascadingselect-parent')
                }

                field_name+='_sub'
                input_type += " cascadingselect-child" 
                child = child.find_element(By.CLASS_NAME, 'cascadingselect-child')

            elif 'aui-field-labelpicker'in child_class :
                input_type += " aui-field-labelpicker"
            elif 'aui-field-datepicker' in child_class :
                input_type += " aui-field-datepicker"

            elif len(auissselects) == 1 :
                input_type += " aui-ss-select"
            elif len(dynamicselects) == 1 :
                input_type += " dynamic-select-mark"
            elif len(jiramultiselects) == 1 :
                input_type += " jira-multi-select"
            elif len(cfselects) == 1 :
                input_type += " cf-select"
            elif len(textareas) == 1 :
                input_type += " textarea"
            elif len(selects) == 1 :
                input_type += " select"
            elif len(texts) == 1 :
                input_type += " text"

        elif 'fieldset' in child.tag_name :
            # find field name 
            input_type += 'fieldset'
            field_name = child.find_element(By.TAG_NAME, 'legend').get_attribute("innerText").split('\n')[0]
            checkbox_elem = child.find_elements(By.TAG_NAME, 'label')
            file_input_elem = child.find_elements(By.CLASS_NAME, 'file-input-list')
            if len(checkbox_elem) > 0 :
                input_type += ' checkbox' 
            
            elif len(file_input_elem) > 0 :
                input_type += ' file-input-list' 
            
            
        elif 'hidden' in child_class:
            input_type += 'hidden'
        else :
            input_type = None

        retval[field_name] = {
            'input_type' : input_type, 
            'required' : required,
            'element' : child
        }
    return retval

def uploading(data, field_infos) :
    for data_item in data[:1] :
        for key in data_item.to_dict().keys() :
            # if not key in ['Project', 'Issue Type'] :
            if key in list(field_infos.keys()):
                element = field_infos[key]['element']
                input_content = data_item[key]
                input_type = field_infos[key]['input_type']
                if input_content :
                    try :
                        match input_type :
                            case 'field-group dynamic-select-mark' :
                                element = element.find_element(By.TAG_NAME, 'select')
                                Select(element).select_by_visible_text(input_content)                        
                            case 'field-group aui-field-cascadingselect cascadingselect-parent' :
                                Select(element).select_by_visible_text(input_content)
                            case 'field-group aui-field-cascadingselect cascadingselect-child' :
                                Select(element).select_by_visible_text(input_content)
                            case 'field-group aui-field-datepicker' :
                                element = element.find_element(By.TAG_NAME, 'input')
                                element.click()
                                element.clear()
                                element.send_keys(Keys.BACK_SPACE)
                                element.send_keys(input_content)
                            case 'field-group cf-select' :
                                # print(key)
                                element = element.find_element(By.TAG_NAME, 'select')
                                Select(element).select_by_visible_text(input_content)
                            case 'field-group select' :
                                select_elements = element.find_elements(By.TAG_NAME, 'select')
                                input_elements = element.find_elements(By.TAG_NAME, 'input')
                                if len(input_elements) == 1 :
                                    element = input_elements[0]                                
                                    element.click()
                                    element.clear()
                                    element.send_keys(Keys.BACK_SPACE)
                                    element.send_keys(input_content)
                                    element.send_keys(Keys.ENTER)
                                elif len(select_elements) == 1 :
                                    element = select_elements[0]
                                    Select(element).select_by_visible_text(input_content)
                            case 'field-group jira-multi-select' :
                                element = element.find_element(By.TAG_NAME, 'textarea')
                                for item in input_content.split(', ') :
                                    element.send_keys(item)
                                    element.send_keys(Keys.ENTER)
                            case 'field-group text' :
                                element = element.find_element(By.TAG_NAME, 'input')
                                if 'ajs-dirty-warning-exempt' in element.get_attribute('class') :
                                    element.click()
                                    element.send_keys(Keys.BACK_SPACE)
                                    element.clear()
                                    element.send_keys(input_content)
                                    time.sleep(3)
                                    element.send_keys(Keys.ENTER)
                                else :
                                    element.clear()
                                    element.send_keys(input_content)
                            case 'fieldset checkbox' :
                                field = element.find_elements(By.TAG_NAME, 'div')
                                options = list(map(lambda x : x.text, field))
                                checkbox = field[options.index(input_content)].find_element(By.TAG_NAME, 'label')
                                checkbox.click()
                            case 'field-group textarea' :
                                if key == 'Description' :
                                    input_content = input_content.replace('[로그:  ]', '[로그: {} ]'.format(data_item['log']))
                                element = element.find_element(By.CLASS_NAME, 'textarea')
                                element.click()
                                element.send_keys(Keys.BACK_SPACE)
                                element.clear()
                                element.send_keys(input_content)
                                time.sleep(3)
                                element.send_keys(Keys.ENTER)                                                    
                            case 'field-group aui-field-labelpicker' :
                                element = element.find_element(By.TAG_NAME, 'textarea')
                                for item in input_content.split(', ') :
                                    element.send_keys(item)
                                    element.send_keys(Keys.ENTER)
                                pass
                            case 'fieldset file-input-list' :
                                filepath = 'C:\\Users\\HAEA_MI\\Downloads\\Feb13JXIssueList.txt'
                                # filepath = 
                                element = element.find_element(By.CLASS_NAME, 'issue-drop-zone__file')
                                element.send_keys(filepath)
                                pass 
                            case 'hidden':
                                pass
                    except AttributeError as e :
                        print('{} field does not support input {}'.format(key, input_content))
    time.sleep(10)

def divide_list(lst, num_groups):
    # Calculate the number of items in each group
    group_size, remainder = divmod(len(lst), num_groups)

    # Initialize an empty list of groups
    groups = [[] for _ in range(num_groups)]

    # Add items to each group
    start = 0
    for i in range(num_groups):
        # Calculate the end index for this group
        end = start + group_size + (i < remainder)

        # Add items to this group
        groups[i] = lst[start:end]

        # Update the start index for the next group
        start = end

    return groups

def set_project(driver, project_option, issuetype_option) :

    project_field = driver.find_element(By.CLASS_NAME, 'issue-setup-fields').find_element(By.ID, 'project-field')
    project_field.click()
    project_field.send_keys(Keys.CONTROL + 'a' + Keys.NULL, '')
    project_field.send_keys(project_option)
    project_field.send_keys(Keys.ENTER)

    time.sleep(5)

    issuetype_field = driver.find_element(By.CLASS_NAME, 'issue-setup-fields').find_element(By.ID, 'issuetype-field')
    issuetype_field.click()
    issuetype_field.send_keys(Keys.CONTROL + 'a' + Keys.NULL, '')
    issuetype_field.send_keys(issuetype_option)
    issuetype_field.send_keys(Keys.ENTER)

    time.sleep(5)

def upload_issues(ID, PW, tasks, auto_upload, show_browser, xlsx_file) :
    options = Options()
    if not show_browser :
        options.headless = True
    
    if not auto_upload :
        driver = webdriver.Chrome('./chromedriver', chrome_options=options)
        driver.get("https://mcols.autoever.com/secure/Dashboard.jspa")
        enter_to_mcols(driver, ID, PW)
        press_create_btn(driver)
        
    num_tasks = len(tasks)
    for task_idx, data_item in enumerate(tasks) :
        try :
            if auto_upload :
                options = Options()
                if not show_browser :
                    options.headless = True
                driver = webdriver.Chrome('./chromedriver', chrome_options=options)
                driver.get("https://mcols.autoever.com/secure/Dashboard.jspa")
                enter_to_mcols(driver, ID, PW)
                press_create_btn(driver)

            set_project(driver, data_item['Project'], data_item['Issue Type'])
            field_infos = type_retriever(driver, data_item['Issue Type'])
            for key in data_item.to_dict().keys() :
                if key in ['Epic Link', 'ESTRACT'] :
                    print('{} is not supported.'.format(key))
                elif not key in ['Project', 'Issue Type', 'excel_info'] :
                    if key in list(field_infos.keys()):
                        element = field_infos[key]['element']
                        input_content = data_item[key]
                        input_type = field_infos[key]['input_type']
                        if input_content :
                            try :
                                match input_type :
                                    case 'field-group dynamic-select-mark' :
                                        element = element.find_element(By.TAG_NAME, 'select')
                                        Select(element).select_by_visible_text(input_content)                        
                                    case 'field-group aui-field-cascadingselect cascadingselect-parent' :
                                        Select(element).select_by_visible_text(input_content)
                                    case 'field-group aui-field-cascadingselect cascadingselect-child' :
                                        Select(element).select_by_visible_text(input_content)
                                    case 'field-group aui-field-datepicker' :
                                        element = element.find_element(By.TAG_NAME, 'input')
                                        element.click()
                                        element.clear()
                                        element.send_keys(Keys.BACK_SPACE)
                                        element.send_keys(input_content)
                                    case 'field-group cf-select' :
                                        # print(key)
                                        element = element.find_element(By.TAG_NAME, 'select')
                                        Select(element).select_by_visible_text(input_content)
                                    case 'field-group select' :
                                        select_elements = element.find_elements(By.TAG_NAME, 'select')
                                        input_elements = element.find_elements(By.TAG_NAME, 'input')
                                        if len(input_elements) == 1 :
                                            element = input_elements[0]                                
                                            element.click()
                                            element.clear()
                                            element.send_keys(Keys.BACK_SPACE)
                                            element.send_keys(input_content)
                                            element.send_keys(Keys.ENTER)
                                        elif len(select_elements) == 1 :
                                            element = select_elements[0]
                                            Select(element).select_by_visible_text(input_content)
                                    case 'field-group jira-multi-select' :
                                        element = element.find_element(By.TAG_NAME, 'textarea')
                                        for item in input_content.split(', ') :
                                            element.send_keys(item)
                                            element.send_keys(Keys.ENTER)
                                    case 'field-group text' :
                                        element = element.find_element(By.TAG_NAME, 'input')
                                        if 'ajs-dirty-warning-exempt' in element.get_attribute('class') :
                                            element.click()
                                            element.send_keys(Keys.BACK_SPACE)
                                            element.clear()
                                            element.send_keys(input_content)
                                            time.sleep(3)
                                            element.send_keys(Keys.ENTER)
                                        else :
                                            element.clear()
                                            element.send_keys(input_content)
                                    case 'fieldset checkbox' :
                                        field = element.find_elements(By.TAG_NAME, 'div')
                                        options = list(map(lambda x : x.text, field))
                                        checkbox = field[options.index(input_content)].find_element(By.TAG_NAME, 'label')
                                        checkbox.click()
                                    case 'field-group textarea' :
                                        if key == 'Description' :
                                            input_content = input_content.replace('[로그:  ]', '[로그: {} ]'.format(data_item['log']))
                                        element = element.find_element(By.CLASS_NAME, 'textarea')
                                        element.click()
                                        element.send_keys(Keys.BACK_SPACE)
                                        element.clear()
                                        element.send_keys(input_content)
                                        time.sleep(3)
                                        element.send_keys(Keys.ENTER)                                                    
                                    case 'field-group aui-field-labelpicker' :
                                        element = element.find_elements(By.TAG_NAME, 'textarea')
                                        if len(element) > 0 :
                                            element = element[0]
                                            for item in input_content.split(', ') :
                                                element.send_keys(item)
                                                element.send_keys(Keys.ENTER)
                                        else :
                                            print('{} : parsing error'.format(key))    
                                        
                                    case 'fieldset file-input-list' :
                                        # filepath = 'C:\\Users\\HAEA_MI\\Downloads\\Feb13JXIssueList.txt'
                                        # # filepath = 
                                        # element = element.find_element(By.CLASS_NAME, 'issue-drop-zone__file')
                                        # element.send_keys(filepath)
                                        pass 
                                    case 'hidden':
                                        pass
                            except AttributeError as e :
                                print('{} field does not support input {}'.format(key, input_content))
            
            if 'attachments' in data_item.to_dict().keys() and 'attachment_files' in data_item.to_dict().keys() :
                if data_item['attachments'] != None and data_item['attachment_files'] != None :
                    element = field_infos['Attachment']['element']
                    element = element.find_element(By.CLASS_NAME, 'issue-drop-zone__file')
                    for file in data_item['attachment_files'].split(', ') :
                        filepath = os.path.join(data_item['attachments'], file)
                        element.send_keys(filepath)
                        timeout = 300
                        while True :
                            try :
                                t = driver.find_element(By.ID, "aui-flag-container")
                                if t :
                                    if timeout == 0 :
                                        break
                                time.sleep(1)
                                timeout-=1
                                break
                            except :
                                pass  
                
                    time.sleep(1)

            time.sleep(2)

            if not auto_upload :
                if task_idx != num_tasks-1 :
                    try :
                        driver.switch_to.new_window('tab')
                        time.sleep(2)
                        driver.get("https://mcols.autoever.com/secure/Dashboard.jspa")
                        press_create_btn(driver)
                    except Exception as e :
                        if 'unexpected alert open' in e :
                            print("unexpected alert open : {}".format(e))
                            driver.switch_to.alert.accept()
                            driver.switch_to.new_window('tab')
                            time.sleep(2)
                            driver.get("https://mcols.autoever.com/secure/Dashboard.jspa")
                            press_create_btn(driver)
            else :
                # # press click()
                # press_submit(driver)
                
                # # get Ticket Number
                # issue_page = move_to_issue_page(driver)
                issue_page = "http://"
                column = data_item['excel_info']['excel_column']
                row_list = data_item['excel_info']['excel_row_items']

                update_cell(xlsx_file, 'Sheet1', row_list.index('upload status')+2, column+1, 1)
                update_cell(xlsx_file, 'Sheet1', row_list.index('issue_tag')+2, column+1, issue_page)
                # uploaded excel
                driver.close()
        except Exception as e :
            if 'unexpected alert open' in e :
                print("unexpected alert open : {}".format(e))
                driver.switch_to.alert.accept()
                if auto_upload :
                    driver.close()
                else :
                    driver.switch_to.new_window('tab')
                    time.sleep(2)
                    driver.get("https://mcols.autoever.com/secure/Dashboard.jspa")
                    press_create_btn(driver)
            
    time.sleep(60)

def update_cell(xlsx_file, sheet_name, row, column, value):
    # Load the workbook
    workbook = load_workbook(filename=xlsx_file)
    # Select the worksheet
    worksheet = workbook[sheet_name]
    # Get the cell
    cell = worksheet.cell(row=row, column=column)
    # Update the cell value
    cell.value = value
    # Save the workbook
    workbook.save(filename=xlsx_file)

def upload_multithread(ID, PW, xlsx_file, num_thread, auto_upload=True, show_browser=True, logs=None) :

    chromedriver_autoinstaller.install()

    df = pd.read_excel(xlsx_file, header=0, index_col=0)
    excel_row_items = list(df.axes[0])
    data = list(df[0:100].to_dict('series').values())
    for idx in range(len(data)) :
        data[idx]['excel_info'] = {
            'excel_column' : idx,
            'excel_row_items' : excel_row_items
        }
    data = list(filter(lambda x : str(x['upload status']) == '0', data))

    if logs :
        logs.insert(tk.END, "Starting upload...\n")

        for tasks in divide_list(data, num_thread):
            t = threading.Thread(target=upload_issues, args=(ID, PW, tasks, auto_upload, show_browser, xlsx_file))
            t.start()
    if logs :
        logs.insert(tk.END, "End upload...\n")


if __name__ == "__main__" :
    chromedriver_autoinstaller.install()
    options = Options()
    driver = webdriver.Chrome('./chromedriver', chrome_options=options)
    driver.get("https://mcols.autoever.com/secure/Dashboard.jspa")

    ID = "10896665"
    PW = "A12345678!"
    xlsx_file = 'HMC OEM 내비게이션 개발 협업 (HMCOEM)_03-09-2023-13-29-22.xlsx'

    enter_to_mcols(driver, ID, PW)
    press_create_btn(driver)

    field_infos = type_retriever(driver, issue_type = 'Bug')
    df = pd.read_excel(xlsx_file, header=0, index_col=0)
    data = list(df[0:100].to_dict('series').values())

    uploading(data, field_infos)

